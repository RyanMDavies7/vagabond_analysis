"""
show_leaf_account_data.py
Creates leaf_data_showcase.xlsx with:
  • API_field_catalog  – every selectable Google-Ads field
  • Account_overview   – 30-day totals
  • Campaign_performance – top 50 campaigns
  • Keyword_stats      – top 100 keywords
  • Conversion_actions – all goals + meta
  • Audiences          – example audience data
"""

import os
import pandas as pd
import datetime as dt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from google.protobuf.json_format import MessageToDict

from packages.google_ads_kpi.auth import get_client

# ── config ──────────────────────────────────────────────────────────────
LEAF_CID = os.getenv("CUSTOMER_ID")       # 10-digit, no dashes
OUTPUT   = "leaf_data_showcase.xlsx"

client   = get_client()
ga       = client.get_service("GoogleAdsService")

# ── helpers ─────────────────────────────────────────────────────────────
def row_to_dict(row):
    """Flatten GoogleAdsRow -> dict"""
    return MessageToDict(row._pb, preserving_proto_field_name=True)

def _scalarize(value):
    """Turn lists / dicts into a JSON string so Excel can store it."""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value

def dump_field_catalog(wb):
    """Add a sheet listing every GAQL field"""
    field_service = client.get_service("GoogleAdsFieldService")
    query = """
      SELECT
        name,
        category,
        selectable,
        filterable,
        sortable
    """
    rows = [
        {
            "name": f.name,
            "category": f.category.name,
            "selectable": f.selectable,
            "filterable": f.filterable,
            "sortable": f.sortable,
        }
        for f in field_service.search_google_ads_fields(query=query)
    ]
    ws = wb.create_sheet("API_field_catalog")
    for r in dataframe_to_rows(pd.DataFrame(rows), index=False, header=True):
        ws.append(r)


def run_query_to_sheet(query: str, sheet_name: str, wb):
    rows = [row_to_dict(r) for r in ga.search(customer_id=LEAF_CID, query=query)]
    if not rows:
        return                                     # nothing to write

    df = pd.json_normalize(rows)                  # flatten nested keys
    ws = wb.create_sheet(sheet_name[:31])

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append([_scalarize(v) for v in row])

# ── sample queries to showcase data ─────────────────────────────────────
SAMPLES = {
    "Account_overview": """
      SELECT customer.descriptive_name,
             customer.currency_code,
             segments.date,
             metrics.impressions,
             metrics.clicks,
             metrics.cost_micros,
             metrics.conversions
      FROM customer
      WHERE segments.date DURING LAST_30_DAYS
    """,
    "Campaign_performance": """
      SELECT campaign.name,
             campaign.status,
             metrics.impressions,
             metrics.clicks,
             metrics.cost_micros,
             metrics.conversions
      FROM campaign
      WHERE segments.date DURING LAST_30_DAYS
      ORDER BY metrics.clicks DESC
      LIMIT 50
    """,
    "Keyword_stats": """
      SELECT ad_group_criterion.keyword.text,
             ad_group_criterion.keyword.match_type,
             metrics.impressions,
             metrics.clicks,
             metrics.conversions,
             metrics.average_cpc
      FROM keyword_view
      WHERE segments.date DURING LAST_30_DAYS
        AND ad_group_criterion.status = 'ENABLED'
      ORDER BY metrics.clicks DESC
      LIMIT 100
    """,
    "Conversion_actions": """
      SELECT conversion_action.name,
             conversion_action.type,
             conversion_action.status,
             conversion_action.value_settings.default_value,
             conversion_action.primary_for_goal,
             conversion_action.attribution_model_settings.attribution_model
      FROM conversion_action
    """,
    "Audiences": """
       SELECT
         user_list.id,
         user_list.name,
         user_list.size_for_display,
         user_list.size_for_search,
         ad_group_criterion.status
       FROM ad_group_audience_view
       LIMIT 100
    """,
}

# ── build workbook ──────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)                       # start fresh sheet list

dump_field_catalog(wb)

for name, q in SAMPLES.items():
    print("Running", name)
    run_query_to_sheet(q, name, wb)

# optional front-cover sheet
cover = wb.create_sheet("README", 0)
cover["A1"] = f"Leaf account data showcase for CID {LEAF_CID}"
cover["A2"] = f"Generated {dt.date.today().isoformat()}"

wb.save(OUTPUT)
print(f"✅ wrote {OUTPUT} for account {LEAF_CID}")
